import openpyxl
import json
import os
from concurrent.futures import ThreadPoolExecutor
from openpyxl.utils import range_boundaries
import re
import pandas as pd
<<<<<<< HEAD
import openpyxl

def detect_header(df, max_rows=10, max_cols=10):
    """
    Detect if the sheet has a proper header or not.
    Looks at the first `max_rows` rows and `max_cols` columns to find data.
    
    Returns:
        cleaned_df (pd.DataFrame): DataFrame with correct header handling
    """
    # Drop completely empty rows and columns first
    df = df.dropna(how="all", axis=0).dropna(how="all", axis=1)

    if df.empty:
        return pd.DataFrame()  # return empty if no data at all

    # Look at a subset (first few rows & columns) for detecting headers
    subset = df.iloc[:max_rows, :max_cols]

    # Case 1: If first row has mostly text and not numeric → assume it’s header
    first_row = subset.iloc[0].astype(str)
    text_ratio = sum(first_row.str.strip() != "") / len(first_row)

    if text_ratio > 0.5:  
        # Likely that the first row is a header
        df.columns = df.iloc[0].astype(str)
        df = df.drop(df.index[0])
    else:
        # Otherwise, just assign generic column names
        df.columns = [f"Column_{i}" for i in range(1, len(df.columns) + 1)]

    # Final cleanup again
    df = df.dropna(how="all", axis=0).dropna(how="all", axis=1)

    return df.reset_index(drop=True)


def read_excel_sheets(file_path):
    """
    Reads only visible and non-empty sheets in the Excel file 
    and returns a dictionary {sheet_name: dataframe}.
    """
    # Load workbook with openpyxl to check sheet visibility
    wb = openpyxl.load_workbook(file_path, read_only=False)
    visible_sheets = [sheet.title for sheet in wb.worksheets if sheet.sheet_state == "visible"]

    print(f"Visible sheets found: {visible_sheets}")  # Debugging

    # Read all visible sheets without headers first
    sheets = pd.read_excel(file_path, sheet_name=visible_sheets, header=None)

    cleaned_sheets = {}
    for name, df in sheets.items():
        cleaned_df = detect_header(df)

        if not cleaned_df.empty:
            print(f"Processing sheet: {name} (rows: {len(cleaned_df)})")     # Debugging line
            cleaned_sheets[name] = cleaned_df
        else:
            print(f"Skipping empty sheet: {name}")

    return cleaned_sheets


# import pandas as pd
# import openpyxl
# from openpyxl.styles import Font, PatternFill

# def detect_header(df, max_rows=10, max_cols=10):
#     """
#     Detect if the sheet has a proper header or not.
#     """
#     df = df.dropna(how="all", axis=0).dropna(how="all", axis=1)

#     if df.empty:
#         return pd.DataFrame()

#     subset = df.iloc[:max_rows, :max_cols]
#     first_row = subset.iloc[0].astype(str)
#     text_ratio = sum(first_row.str.strip() != "") / len(first_row)

#     if text_ratio > 0.5:
#         df.columns = df.iloc[0].astype(str)
#         df = df.drop(df.index[0])
#     else:
#         df.columns = [f"Column_{i}" for i in range(1, len(df.columns) + 1)]

#     df = df.dropna(how="all", axis=0).dropna(how="all", axis=1)
#     return df.reset_index(drop=True)


# def read_excel_sheets(file_path):
#     """
#     Reads visible sheets in Excel, extracting both data (pandas df) and styles (dict).
#     Returns: {sheet_name: {"data": df, "styles": styles}}
#     """
#     wb = openpyxl.load_workbook(file_path, read_only=False, data_only=True)
#     visible_sheets = [sheet for sheet in wb.worksheets if sheet.sheet_state == "visible"]

#     result = {}

#     for sheet in visible_sheets:
#         # --- Extract raw values into dataframe ---
#         data = sheet.values
#         df = pd.DataFrame(data)
#         cleaned_df = detect_header(df)

#         if cleaned_df.empty:
#             continue

#         # --- Extract styles ---
#         styles = {}
#         for row in sheet.iter_rows():
#             for cell in row:
#                 if cell.row == 1:  # skip column headers row before cleaning
#                     continue
#                 if cell.value is None:
#                     continue
#                 styles[(cell.row - 1, cell.column - 1)] = {
#                     "font_name": cell.font.name,
#                     "font_size": cell.font.sz,
#                     "bold": cell.font.bold,
#                     "italic": cell.font.italic,
#                     "color": (cell.font.color.rgb if cell.font.color else "000000"),
#                     "fill_color": (cell.fill.fgColor.rgb if isinstance(cell.fill, PatternFill) else None),
#                 }

#         result[sheet.title] = {
#             "data": cleaned_df,
#             "styles": styles
#         }

#     return result
=======
from datetime import datetime

def get_fill_color(cell):
    fill = cell.fill
    if fill is None or fill.patternType is None or fill.patternType == 'none':
        return None  # No fill
    # Direct RGB, e.g. 'FFB6DDE8' (ARGB: alpha+RGB)
    if hasattr(fill, 'fgColor'):
        color = fill.fgColor
        # If color is RGB (most direct case)
        if color.type == 'rgb' and color.rgb:
            return color.rgb  # format: 'FFB6DDE8'
        # Try theme-based colors (fallback, not always accurate)
        elif color.type == 'theme' and hasattr(color, 'rgb') and color.rgb:
            return color.rgb
        elif color.type == 'indexed' and color.indexed is not None:
            # Map Excel's indexed to RGB:
            from openpyxl.styles.colors import COLOR_INDEX
            rgb = COLOR_INDEX.get(color.indexed)
            return rgb
    return None

def extract_style(cell, style_cache):
    """
    Extracts and caches style info for a cell.
    Returns a dict with font/fill/alignment details.
    """
    if not cell.has_style:
        return None

    # Build a key to cache repeated styles
    key = (
        cell.font.bold, cell.font.italic, cell.font.underline,
        # getattr(cell.font.color, "rgb", None),
        # getattr(cell.fill.fgColor, "rgb", None),
        get_fill_color(cell),
        
        cell.number_format, cell.alignment.horizontal, cell.alignment.vertical
    )

    if key not in style_cache:
        style_cache[key] = {
            "font_bold": cell.font.bold,
            "font_italic": cell.font.italic,
            "font_underline": cell.font.underline,
            # "font_color": getattr(cell.font.color, "rgb", None),
            # "fill_color": getattr(cell.fill.fgColor, "rgb", None),
            "font_color": get_fill_color(cell),
            "fill_color": get_fill_color(cell),
            "number_format": cell.number_format,
            "alignment_horizontal": cell.alignment.horizontal,
            "alignment_vertical": cell.alignment.vertical,
        }

    return style_cache[key]

def expand_merged_cells(ws):
    """
    Expand merged cells into individual cells by copying
    the top-left cell's value into all merged cells.
    """
    merged_ranges = list(ws.merged_cells.ranges)  # copy because we’ll unmerge

    for merged_range in merged_ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        top_left_cell = ws.cell(row=min_row, column=min_col)
        top_left_value = top_left_cell.value

        # Unmerge before assigning values
        ws.unmerge_cells(str(merged_range))

        # Fill each cell in range with top-left value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col, value=top_left_value)
                
def filter_latest_comments(text, max_comments=3):
    """
    Extracts only the latest max_comments dated comments from a multi-line Comments field.
    """
    if pd.isna(text) or text in (None, ""):
        return text
    
    comments = str(text).split("\n")
    date_map = {}
    for c in comments:
        dates = re.findall(r"\b\d{2}/\d{2}\b", c)
        if dates:
            for d in dates:
                date_map.setdefault(d, []).append(c)
    
    if not date_map:
        return text
    
    # Sort dates descending by MM/DD: month as int then day as int
    sorted_dates = sorted(
        date_map.keys(),
        key=lambda d: (int(d.split("/")[0]), int(d.split("/")[1])),
        reverse=True,
    )
    
    # Select top N dates only
    top_dates = sorted_dates[:max_comments]
    
    # Collect comments for these top dates
    selected_comments = []
    for d in top_dates:
        selected_comments.extend(date_map[d])
    
    # Remove duplicate comments if any while preserving order
    seen = set()
    unique_comments = []
    for c in selected_comments:
        if c not in seen:
            unique_comments.append(c)
            seen.add(c)
    
    return "\n".join(unique_comments)



def process_sheet(ws, output_folder, include_styles=True, stream=True):
    """
    Process a single worksheet: find headers, extract rows, and save JSON.
    """
    
    #expand merged cells first
    expand_merged_cells(ws)
    
    sheet_name = ws.title
    style_cache = {}

    # 🔎 Find header row
    header_row_idx = None
    headers = []
    for i in range(1, ws.max_row + 1):
        values = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[i]]
        valid_headers = [h for h in values if h and not h.lower().startswith("unnamed")]
        if len(valid_headers) >= 2:
            header_row_idx = i
            headers = values
            break

    if not headers:
        print(f"Sheet '{sheet_name}' skipped: header not found.")
        return None

    # 🔎 Clean headers
    keep_indices = [
        idx for idx, h in enumerate(headers)
        if h and not h.lower().startswith("unnamed")
    ]
    final_headers = [headers[idx] for idx in keep_indices]
    print(f"Extracting '{sheet_name}': header row {header_row_idx}: {final_headers}")

    # 🔎 JSON path
    json_path = os.path.join(output_folder, f"{sheet_name}.json")

    if stream:
        # ✅ Stream JSON row by row (low memory)
        with open(json_path, "w", encoding="utf-8") as f:
            f.write("[\n")
            first = True
            for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row, max_col=len(headers)):
                row_data = {}
                for pos, idx in enumerate(keep_indices):
                    cell = row[idx]
                    value = cell.value
                    # Convert datetime to date only if time not required
                    if isinstance(value, datetime):
                       value = value.date()
                    row_data[final_headers[pos]] = {
                        "value": cell.value,
                        "style": extract_style(cell, style_cache) if include_styles else None
                    }

                # Skip empty rows
                if any(row_data[header]["value"] not in (None, "") for header in final_headers):
                    json_str = json.dumps(row_data, default=str, indent=2)
                    if not first:
                        f.write(",\n")
                    first = False
                    f.write(json_str)
            f.write("\n]")
    else:
        # ✅ Collect all rows in memory (good for small data)
        all_rows = []
        for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row, max_col=len(headers)):
            row_data = {}
            for pos, idx in enumerate(keep_indices):
                cell = row[idx]
                value = cell.value
                
                # Convert datetime to date only if time not required
                if isinstance(value, datetime):
                    value = value.date()
                
                # Handle Comments column → keep only latest
                if final_headers[pos].lower() == "comments":
                   value = filter_latest_comments(value)

                row_data[final_headers[pos]] = {
                    "value": cell.value,
                    "style": extract_style(cell, style_cache) if include_styles else None
                }
            if any(row_data[header]["value"] not in (None, "") for header in final_headers):
                json_str = json.dumps(row_data, default=str, indent=2)
                if not first:
                    f.write(",\n")
                f.write(json_str)
                first = False
                f.write("\n]")
                all_rows.append(row_data)

        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(all_rows, f, indent=2, default=str)

    print(f"✅ Saved JSON for sheet '{sheet_name}' → {json_path}")
    return json_path


def read_visible_excel_sheets_to_json(file_path, output_folder="extracted_json", include_styles=True, stream=True, sheets=None, parallel=False):
    """
    Convert Excel sheets to JSON (scalable).
    - include_styles: whether to extract formatting info
    - stream: write JSON row by row (low memory)
    - sheets: list of sheet names to include (None = all visible)
    - parallel: process multiple sheets in parallel
    """
    os.makedirs(output_folder, exist_ok=True)

    # Use read_only if not extracting styles (faster for large files)
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # Select sheets
    visible_sheets = [ws for ws in wb.worksheets if ws.sheet_state == "visible"]
    if sheets:
        visible_sheets = [ws for ws in visible_sheets if ws.title in sheets]

    saved_files = []

    if parallel:
        with ThreadPoolExecutor() as executor:
            futures = [executor.submit(process_sheet, ws, output_folder, include_styles, stream) for ws in visible_sheets]
            saved_files = [f.result() for f in futures if f.result()]
    else:
        for ws in visible_sheets:
            result = process_sheet(ws, output_folder, include_styles, stream)
            if result:
                saved_files.append(result)

    return saved_files
>>>>>>> origin/Priyanshi
