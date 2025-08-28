# from flask import Flask, request, jsonify, send_file
# from flask_cors import CORS
# import os
# import uuid
# import shutil
 
# app = Flask(__name__)
# CORS(app)  # Enable CORS for all routes
 
# UPLOAD_FOLDER = "./uploads"
# GENERATED_FOLDER = "./generated"
 
# os.makedirs(UPLOAD_FOLDER, exist_ok=True)
# os.makedirs(GENERATED_FOLDER, exist_ok=True)
 
 
# @app.route("/api/upload-report", methods=["POST"])
# def upload_report():
#     # Check both files in request
#     if 'excel' not in request.files or 'ppt' not in request.files:
#         return jsonify({"message": "Both Excel and PPT files are required"}), 400
 
#     excel_file = request.files['excel']
#     ppt_file = request.files['ppt']
 
#     if excel_file.filename == '' or ppt_file.filename == '':
#         return jsonify({"message": "Invalid file(s)"}), 400
 
#     # Generate unique ID for this request
#     unique_id = str(uuid.uuid4())
 
#     # Save Excel
#     excel_upload_path = os.path.join(UPLOAD_FOLDER, f"{unique_id}.xlsx")
#     excel_file.save(excel_upload_path)
 
#     # Save PPT
#     ppt_upload_path = os.path.join(UPLOAD_FOLDER, f"{unique_id}.pptx")
#     ppt_file.save(ppt_upload_path)
 
#     # PROCESSING LOGIC GOES HERE
#     # For demo, just copy them as "generated"
#     excel_generated_path = os.path.join(GENERATED_FOLDER, f"{unique_id}_report.xlsx")
#     ppt_generated_path = os.path.join(GENERATED_FOLDER, f"{unique_id}_report.pptx")
 
#     shutil.copyfile(excel_upload_path, excel_generated_path)
#     shutil.copyfile(ppt_upload_path, ppt_generated_path)
 
#     # Return download URLs for both
#     response = {
#         "excelDownloadUrl": f"http://localhost:5000/api/download-report/{unique_id}/excel",
#         "pptDownloadUrl": f"http://localhost:5000/api/download-report/{unique_id}/ppt"
#     }
 
#     return jsonify(response)
 
 
# @app.route("/api/download-report/<file_id>/<file_type>", methods=["GET"])
# def download_report(file_id, file_type):
#     if file_type == "excel":
#         file_path = os.path.join(GENERATED_FOLDER, f"{file_id}_report.xlsx")
#         download_name = "report.xlsx"
#         mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#     elif file_type == "ppt":
#         file_path = os.path.join(GENERATED_FOLDER, f"{file_id}_report.pptx")
#         download_name = "report.pptx"
#         mimetype = "application/vnd.openxmlformats-officedocument.presentationml.presentation"
#     else:
#         return jsonify({"message": "Invalid file type"}), 400
 
#     if not os.path.exists(file_path):
#         return jsonify({"message": "Report not found"}), 404
 
#     return send_file(file_path,
#                      as_attachment=True,
#                      download_name=download_name,
#                      mimetype=mimetype)
 
 
# if __name__ == "__main__":
#     app.run(debug=True)

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import os
import uuid
import shutil
from openpyxl import load_workbook  # Import openpyxl

app = Flask(__name__)
CORS(app)

UPLOAD_FOLDER = "./uploads"
GENERATED_FOLDER = "./generated"

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(GENERATED_FOLDER, exist_ok=True)

@app.route("/api/upload-report", methods=["POST"])
def upload_report():
    if 'excel' not in request.files or 'ppt' not in request.files:
        return jsonify({"message": "Both Excel and PPT files are required"}), 400

    excel_file = request.files['excel']
    ppt_file = request.files['ppt']

    if excel_file.filename == '' or ppt_file.filename == '':
        return jsonify({"message": "Invalid file(s)"}), 400

    unique_id = str(uuid.uuid4())

    # Save Excel
    excel_upload_path = os.path.join(UPLOAD_FOLDER, f"{unique_id}.xlsx")
    excel_file.save(excel_upload_path)

    # Save PPT
    ppt_upload_path = os.path.join(UPLOAD_FOLDER, f"{unique_id}.pptx")
    ppt_file.save(ppt_upload_path)

    # Extract sheet names from the uploaded Excel file
    workbook = load_workbook(excel_upload_path, read_only=True)
    sheet_names = workbook.sheetnames  # Get the list of sheet names

    # PROCESSING LOGIC GOES HERE - for demo, just copy files as generated
    excel_generated_path = os.path.join(GENERATED_FOLDER, f"{unique_id}_report.xlsx")
    ppt_generated_path = os.path.join(GENERATED_FOLDER, f"{unique_id}_report.pptx")

    shutil.copyfile(excel_upload_path, excel_generated_path)
    shutil.copyfile(ppt_upload_path, ppt_generated_path)

    response = {
        "excelDownloadUrl": f"http://localhost:5000/api/download-report/{unique_id}/excel",
        "pptDownloadUrl": f"http://localhost:5000/api/download-report/{unique_id}/ppt",
        "sheetNames": sheet_names  # Return extracted sheet names
    }

    return jsonify(response)


@app.route("/api/download-report/<file_id>/<file_type>", methods=["GET"])
def download_report(file_id, file_type):
    if file_type == "excel":
        file_path = os.path.join(GENERATED_FOLDER, f"{file_id}_report.xlsx")
        download_name = "report.xlsx"
        mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif file_type == "ppt":
        file_path = os.path.join(GENERATED_FOLDER, f"{file_id}_report.pptx")
        download_name = "report.pptx"
        mimetype = "application/vnd.openxmlformats-officedocument.presentationml.presentation"
    else:
        return jsonify({"message": "Invalid file type"}), 400

    if not os.path.exists(file_path):
        return jsonify({"message": "Report not found"}), 404

    return send_file(file_path,
                     as_attachment=True,
                     download_name=download_name,
                     mimetype=mimetype)


if __name__ == "__main__":
    app.run(debug=True)
 